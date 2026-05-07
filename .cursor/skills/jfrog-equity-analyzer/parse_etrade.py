#!/usr/bin/env python3
"""
Parse an eTrade ByBenefitType_expanded.xlsx file into structured JSON.

Usage:
    python3 parse_etrade.py <xlsx_path>

Output: JSON object with keys `rsus`, `options`, `espp`. Dates are ISO strings.
Requires: pip install openpyxl
"""
import json
import re
import sys
from datetime import datetime, date
from typing import Any

try:
    import openpyxl  # type: ignore
except ImportError:
    sys.stderr.write("Missing dependency: pip install openpyxl\n")
    sys.exit(1)


MONTHS = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12,
}


def parse_date(value: Any) -> str | None:
    if value is None or value == '' or value == 'NA':
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if not isinstance(value, str):
        return None
    s = value.strip()
    if not s or s == 'NA' or s == '—':
        return None
    m1 = re.match(r'^(\d{1,2})-([A-Z]{3})-(\d{4})$', s, re.IGNORECASE)
    if m1:
        month = MONTHS.get(m1.group(2).upper())
        if month:
            return date(int(m1.group(3)), month, int(m1.group(1))).isoformat()
    m2 = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m2:
        return date(int(m2.group(3)), int(m2.group(1)), int(m2.group(2))).isoformat()
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        return None


def parse_dollar(value: Any) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return 0.0
    cleaned = re.sub(r'[$,\s]', '', value)
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_pct(value: Any) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) / 100 if value > 1 else float(value)
    if not isinstance(value, str):
        return 0.0
    cleaned = re.sub(r'[%\s]', '', value)
    try:
        n = float(cleaned)
    except ValueError:
        return 0.0
    return n / 100 if n > 1 else n


def num(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(',', ''))
        except ValueError:
            return 0.0
    return 0.0


def s(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def parse_espp(rows: list[tuple]) -> list[dict]:
    out = []
    for r in rows[1:]:
        if s(r[0]) != 'Purchase':
            continue
        purchase_date = parse_date(r[2])
        grant_date = parse_date(r[10])
        if not purchase_date:
            continue
        out.append({
            'purchaseDate': purchase_date,
            'grantDate': grant_date or purchase_date,
            'purchasePrice': num(r[3]),
            'purchasedQty': num(r[4]),
            'blockedQty': num(r[12]),
            'discountPct': parse_pct(r[15]),
            'grantDateFmv': parse_dollar(r[16]),
            'purchaseDateFmv': parse_dollar(r[17]),
        })
    return out


def parse_rsu(rows: list[tuple]) -> list[dict]:
    grants: list[dict] = []
    current: dict | None = None
    for r in rows[1:]:
        rt = s(r[0])
        if rt == 'Grant':
            grant_date = parse_date(r[2])
            if not grant_date:
                continue
            current = {
                'grantNumber': s(r[11]),
                'grantDate': grant_date,
                'fmvAtGrant': 0.0,
                'grantedQty': num(r[4]),
                'vestedQty': num(r[6]),
                'unvestedQty': num(r[7]),
                'blockedQty': num(r[14]),
                'vestSchedule': [],
            }
            grants.append(current)
        elif rt == 'Vest Schedule' and current is not None:
            vest_date = parse_date(r[19])
            if not vest_date:
                continue
            current['vestSchedule'].append({
                'vestDate': vest_date,
                'qty': num(r[21]),
                'fmvAtVest': num(r[57]) or None,
            })

    for g in grants:
        if not g['vestSchedule']:
            continue
        sorted_v = sorted(g['vestSchedule'], key=lambda v: v['vestDate'])
        first = next((v for v in sorted_v if v.get('fmvAtVest')), None)
        if first:
            g['fmvAtGrant'] = first['fmvAtVest']
    return grants


def parse_options(rows: list[tuple]) -> list[dict]:
    grants: list[dict] = []
    current: dict | None = None
    for r in rows[1:]:
        rt = s(r[0])
        if rt == 'Grant':
            grant_date = parse_date(r[2])
            if not grant_date:
                continue
            current = {
                'grantNumber': s(r[9]),
                'grantDate': grant_date,
                'exercisePrice': num(r[4]),
                'grantedQty': num(r[3]),
                'exercisableQty': num(r[6]),
                'blockedQty': num(r[14]) or num(r[15]),
                'expirationDate': None,
                'vestSchedule': [],
            }
            grants.append(current)
        elif rt == 'Vest Schedule' and current is not None:
            vest_date = parse_date(r[19])
            if not vest_date:
                continue
            exp = parse_date(r[27])
            if exp and not current['expirationDate']:
                current['expirationDate'] = exp
            current['vestSchedule'].append({
                'vestDate': vest_date,
                'qty': num(r[20]),
            })
    return grants


def main() -> None:
    if len(sys.argv) < 2:
        sys.stderr.write('Usage: parse_etrade.py <xlsx_path>\n')
        sys.exit(1)
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def sheet_rows(*names: str) -> list[tuple]:
        for n in names:
            if n in wb.sheetnames:
                ws = wb[n]
                return list(ws.iter_rows(values_only=True))
        return []

    result = {
        'espp': parse_espp(sheet_rows('ESPP')),
        'rsus': parse_rsu(sheet_rows('Restricted Stock', 'Restricted Stock & Performance')),
        'options': parse_options(sheet_rows('Options', 'Stock Options')),
    }
    json.dump(result, sys.stdout, indent=2, ensure_ascii=False)
    sys.stdout.write('\n')


if __name__ == '__main__':
    main()
