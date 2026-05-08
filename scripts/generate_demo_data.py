#!/usr/bin/env python3
"""
Generate a demo eTrade ByBenefitType_expanded.xlsx for illustration purposes.

Data is entirely fictional with round numbers.
Targets (at FROG=$50 default app price, USD/ILS=3.6, salary=300K NIS):
  Available now  ≈ 152,000 ILS net after tax
    RSU blocked    600 shares  × $35.25 net × 3.6  =  76,140 ILS  (capital track)
    Options ITM    500 opts    × $22.50 net × 3.6  =  40,500 ILS  (strike $20)
    ESPP P1        150 shares  × $40.65 net × 3.6  =  21,951 ILS  (capital)
    ESPP P2        100 shares  × $37.20 net × 3.6  =  13,392 ILS  (ordinary)
                                                    ─────────────
                                                       151,983 ILS

  Future vesting ≈ 102,000 ILS net (RSU only, FutureVestsTable)
    RSU-2021 future  2 × 100 shares  (capital)
    RSU-2024 future  4 × 125 shares  (mixed ordinary/capital)
                                                    ─────────────
                                                       ≈102,000 ILS

Usage:
    pip install openpyxl
    python3 scripts/generate_demo_data.py
"""
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise SystemExit("Missing dependency: pip install openpyxl")


def _row(n: int = 60) -> list:
    return [""] * n


# ─── RSU sheet ────────────────────────────────────────────────────────────────

def _rsu_grant(gnum, gdate, granted, vested, unvested, blocked) -> list:
    r = _row()
    r[0] = "Grant"
    r[2] = gdate
    r[4] = granted
    r[6] = vested
    r[7] = unvested
    r[11] = gnum
    r[14] = blocked
    return r


def _rsu_vest(vdate, qty, fmv=None) -> list:
    r = _row()
    r[0] = "Vest Schedule"
    r[19] = vdate
    r[21] = qty
    if fmv is not None:
        r[57] = fmv
    return r


def build_rsu_sheet(ws) -> None:
    # Header row (parser skips rows[0])
    h = _row()
    h[0] = "Row Type"; h[2] = "Date"; h[4] = "Shares Granted"
    h[6] = "Shares Vested"; h[7] = "Shares Unvested"
    h[11] = "Grant Number"; h[14] = "Shares Blocked"
    h[19] = "Vest Date"; h[21] = "Vest Qty"; h[57] = "FMV at Vest"
    ws.append(h)

    # ── Grant 1 ──────────────────────────────────────────────────────────────
    # RSU-2021-001  granted 2021-03-01  900 total
    # 700 vested (200+200+200+100), 200 unvested, 600 blocked
    # At $50 default price: 600 × $35.25 net × 3.6 ≈ ₪76,140 available now
    ws.append(_rsu_grant("RSU-2021-001", date(2021, 3, 1),
                          granted=900, vested=700, unvested=200, blocked=600))
    ws.append(_rsu_vest(date(2022, 3,  1), 200, fmv=15.00))   # past
    ws.append(_rsu_vest(date(2023, 3,  1), 200, fmv=17.00))   # past
    ws.append(_rsu_vest(date(2024, 3,  1), 200, fmv=20.00))   # past
    ws.append(_rsu_vest(date(2025, 3,  1), 100, fmv=22.00))   # past
    ws.append(_rsu_vest(date(2026, 9,  1), 100))               # future (~4 months)
    ws.append(_rsu_vest(date(2027, 3,  1), 100))               # future (~10 months)

    # ── Grant 2 ──────────────────────────────────────────────────────────────
    # RSU-2024-001  granted 2024-01-15  500 total
    # 0 vested, 500 unvested, 0 blocked
    # At $50: 500 × (mix ordinary/capital) × 3.6 ≈ ₪66,375 future
    ws.append(_rsu_grant("RSU-2024-001", date(2024, 1, 15),
                          granted=600, vested=0, unvested=600, blocked=0))
    ws.append(_rsu_vest(date(2026,  7, 15), 150))   # future – ordinary track
    ws.append(_rsu_vest(date(2027,  1, 15), 150))   # future – capital track
    ws.append(_rsu_vest(date(2028,  1, 15), 150))   # future – capital track
    ws.append(_rsu_vest(date(2029,  1, 15), 150))   # future – capital track


# ─── Options sheet ────────────────────────────────────────────────────────────

def _opt_grant(gnum, gdate, granted, strike, exercisable, blocked=0) -> list:
    r = _row()
    r[0] = "Grant"
    r[2] = gdate
    r[3] = granted
    r[4] = strike
    r[6] = exercisable
    r[9] = gnum
    r[14] = blocked
    return r


def _opt_vest(vdate, qty, exp=None) -> list:
    r = _row()
    r[0] = "Vest Schedule"
    r[19] = vdate
    r[20] = qty
    if exp:
        r[27] = exp
    return r


def build_options_sheet(ws) -> None:
    h = _row()
    h[0] = "Row Type"; h[2] = "Date"; h[3] = "Shares Granted"
    h[4] = "Exercise Price"; h[6] = "Exercisable"
    h[9] = "Grant Number"; h[14] = "Blocked"
    h[19] = "Vest Date"; h[20] = "Vest Qty"; h[27] = "Expiration Date"
    ws.append(h)

    exp = date(2031, 6, 1)

    # ── Grant 1 ──────────────────────────────────────────────────────────────
    # OPT-2019-001  granted 2019-06-01  2,000 options at $20 strike
    # 500 exercisable (4 past vests × 125), 1,500 unvested (3 future × 500)
    # At $50: 500 × ($50-$20) × 0.75 × 3.6 = ₪40,500 available now
    ws.append(_opt_grant("OPT-2019-001", date(2019, 6, 1),
                          granted=2000, strike=20.00, exercisable=500))
    ws.append(_opt_vest(date(2020, 6,  1), 125, exp))   # past
    ws.append(_opt_vest(date(2021, 6,  1), 125, exp))   # past
    ws.append(_opt_vest(date(2022, 6,  1), 125, exp))   # past
    ws.append(_opt_vest(date(2023, 6,  1), 125, exp))   # past
    ws.append(_opt_vest(date(2026, 12, 1), 500, exp))   # future (~7 months)
    ws.append(_opt_vest(date(2027, 6,  1), 500, exp))   # future (~13 months)
    ws.append(_opt_vest(date(2028, 6,  1), 500, exp))   # future (~25 months)


# ─── ESPP sheet ───────────────────────────────────────────────────────────────

def _espp_purchase(pdate, gdate, price, qty, blocked, discount,
                   grant_fmv, purchase_fmv) -> list:
    r = _row()
    r[0] = "Purchase"
    r[2] = pdate
    r[3] = price
    r[4] = qty
    r[10] = gdate
    r[12] = blocked
    r[15] = discount
    r[16] = grant_fmv
    r[17] = purchase_fmv
    return r


def build_espp_sheet(ws) -> None:
    h = _row()
    h[0] = "Row Type"; h[2] = "Purchase Date"; h[3] = "Purchase Price"
    h[4] = "Purchased Qty"; h[10] = "Grant Date"; h[12] = "Blocked Qty"
    h[15] = "Discount %"; h[16] = "Grant Date FMV"; h[17] = "Purchase Date FMV"
    ws.append(h)

    # ── Purchase 1 ───────────────────────────────────────────────────────────
    # Enrollment 2023-04-01 → 2023-09-30  (capital track – qualifying period passed)
    # Grant FMV $18, Purchase FMV $19  →  price = 0.85 × $18 = $15.30 → $15
    # 150 shares purchased, all blocked
    # At $50: 150 × $40.65 net × 3.6 ≈ ₪21,951 available now
    ws.append(_espp_purchase(
        pdate=date(2023, 9, 30), gdate=date(2023, 4, 1),
        price=15.00, qty=150, blocked=150,
        discount=0.15, grant_fmv=18.00, purchase_fmv=19.00,
    ))

    # ── Purchase 2 ───────────────────────────────────────────────────────────
    # Enrollment 2024-04-01 → 2024-09-30  (ordinary track – not yet 24 months)
    # Grant FMV $21, Purchase FMV $22  →  price = 0.85 × $21 = $17.85 → $18
    # 100 shares purchased, all blocked
    # At $50: 100 × $37.20 net × 3.6 ≈ ₪13,392 available now
    ws.append(_espp_purchase(
        pdate=date(2024, 9, 30), gdate=date(2024, 4, 1),
        price=18.00, qty=100, blocked=100,
        discount=0.15, grant_fmv=21.00, purchase_fmv=22.00,
    ))


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    build_rsu_sheet(wb.create_sheet("Restricted Stock"))
    build_options_sheet(wb.create_sheet("Stock Options"))
    build_espp_sheet(wb.create_sheet("ESPP"))

    out = Path(__file__).parent.parent / "sample_data" / "demo_ByBenefitType_expanded.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
